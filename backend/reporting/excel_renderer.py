"""Excel renderer over the section payload (unify-report-format-coverage,
phase 2).

Each section becomes one worksheet named after its title. Blocks are written
top to bottom: a StatGrid as a labelled KPI block, a Table as a header + rows,
a Narrative as a bold heading plus wrapped paragraphs, a Meter as a single
percentage cell.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.reporting.section_data import (
    Block,
    Meter,
    Narrative,
    SectionData,
    StatGrid,
    Table,
)

SUPPORTED_BLOCKS: frozenset[type] = frozenset({StatGrid, Table, Narrative, Meter})

_INVALID_SHEET_CHARS = set(r"[]:*?/\\")
_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _safe_sheet_title(title: str, wb: Workbook) -> str:
    cleaned = "".join(" " if ch in _INVALID_SHEET_CHARS else ch for ch in title).strip()
    cleaned = (cleaned or "Section")[:31]
    # Excel sheet names must be unique; suffix if the base is taken.
    base, candidate, n = cleaned, cleaned, 2
    while candidate in wb.sheetnames:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    return candidate


def _write_stat_grid(ws: Worksheet, block: StatGrid, row: int) -> int:
    for item in block.items:
        ws.cell(row=row, column=1, value=item.label).font = _HEADER_FONT
        ws.cell(row=row, column=2, value=item.value)
        if item.sub:
            ws.cell(row=row, column=3, value=item.sub)
        row += 1
    return row + 1


def _write_table(ws: Worksheet, block: Table, row: int) -> int:
    for col_idx, name in enumerate(block.columns, start=1):
        ws.cell(row=row, column=col_idx, value=name).font = _HEADER_FONT
    row += 1
    for data_row in block.rows:
        for col_idx, cell in enumerate(data_row, start=1):
            ws.cell(row=row, column=col_idx, value=cell)
        row += 1
    return row + 1


def _write_narrative(ws: Worksheet, block: Narrative, row: int) -> int:
    ws.cell(row=row, column=1, value=block.heading).font = _HEADER_FONT
    row += 1
    for paragraph in block.paragraphs:
        cell = ws.cell(row=row, column=1, value=paragraph)
        cell.alignment = _WRAP
        row += 1
    return row + 1


def _write_meter(ws: Worksheet, block: Meter, row: int) -> int:
    pct = max(0, min(100, round(block.pct)))
    ws.cell(row=row, column=1, value=block.label).font = _HEADER_FONT
    ws.cell(row=row, column=2, value=f"{pct}%")
    return row + 2


def _write_block(ws: Worksheet, block: Block, row: int) -> int:
    if isinstance(block, StatGrid):
        return _write_stat_grid(ws, block, row)
    if isinstance(block, Table):
        return _write_table(ws, block, row)
    if isinstance(block, Narrative):
        return _write_narrative(ws, block, row)
    if isinstance(block, Meter):
        return _write_meter(ws, block, row)
    raise TypeError(f"Excel renderer cannot render {type(block).__name__}")


def render_excel(section: SectionData, wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(_safe_sheet_title(section.title, wb))
    row = 1
    for block in section.blocks:
        row = _write_block(ws, block, row)
    return ws
