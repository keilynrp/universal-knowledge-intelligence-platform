"""The presentation contract in a workbook (report-presentation 5.2, 5.3).

Excel is not a document, so it gets a *placement* rather than a rendering: a
dedicated `Methodology` sheet carrying every rendered section's finding and
disclosure, and a caveat row above each table so a copied range keeps its
warning. That second one is the point of the whole Excel branch — this is the
format most often re-cut and pasted into a deck, which is exactly the path by
which a proxy metric loses the sentence saying it is a proxy.
"""
import io

import openpyxl

from backend import models
from backend.exporters.excel_exporter import EnterpriseExcelExporter
import pytest

pytestmark = pytest.mark.reporting

_SECTIONS = ["entity_stats", "journal_portfolio", "harmonization_log"]


def _seed(db) -> None:
    for idx in range(4):
        db.add(models.RawEntity(
            primary_label=f"Record {idx}", domain="default",
            validation_status="valid" if idx else "pending",
            enrichment_status="completed",
            enrichment_concepts="knowledge graph; ontology",
            enrichment_citation_count=100 + idx,
            enrichment_source="openalex",
            secondary_label="Review",
            quality_score=0.8,
        ))
    db.add(models.HarmonizationLog(
        step_id="normalize_labels", step_name="Normalize labels",
        records_updated=4, fields_modified="primary_label",
    ))
    # A journal so the NIF caveat — the one this contract exists for — is real.
    db.add(models.JournalMetric(
        org_id=None, issn_l="issn-x", display_name="Nature Methods",
        normalized_impact_factor=4.10, nif_field="cs",
        nif_bayes=4.05, nif_ci_low=3.60, nif_ci_high=4.55,
        works_2yr=8, apc_usd=1500, is_in_doaj=True,
    ))
    db.commit()


def _seed_singular(db) -> None:
    """One of everything — the only dataset where plural agreement can go wrong."""
    db.add(models.RawEntity(
        primary_label="Only", domain="default", validation_status="valid",
        enrichment_status="completed", enrichment_concepts="ontology",
        enrichment_citation_count=10, enrichment_source="openalex",
        secondary_label="Review", quality_score=0.8,
    ))
    db.add(models.HarmonizationLog(
        step_id="normalize_labels", step_name="Normalize labels",
        records_updated=1, fields_modified="primary_label",
    ))
    db.add(models.AuthorityRecord(
        field_name="brand_capitalized", original_value="acme corp",
        canonical_label="ACME Corporation", confidence=0.92, status="pending",
        resolution_status="ambiguous", review_required=True,
    ))
    db.add(models.Author(id=1, name_key="alice", display_name="Alice Ng"))
    db.add(models.AuthorStats(
        author_id=1, org_id=None, domain_id="default",
        degree=1, centrality=0.5, community_id=1, publication_count=1,
    ))
    db.add(models.JournalMetric(
        org_id=None, issn_l="issn-x", display_name="Nature Methods",
        normalized_impact_factor=4.10, nif_field="cs", nif_bayes=4.05,
        nif_ci_low=3.60, nif_ci_high=4.55, works_2yr=1, apc_usd=1500,
        is_in_doaj=True,
    ))
    db.commit()


def _workbook(db, sections=None):
    data = EnterpriseExcelExporter().build(db, "default", sections or _SECTIONS)
    return openpyxl.load_workbook(io.BytesIO(data))


def _rows(ws) -> list[tuple]:
    return [
        tuple("" if c.value is None else str(c.value) for c in row)
        for row in ws.iter_rows()
    ]


# ── 5.2 Methodology sheet ─────────────────────────────────────────────────────


def test_workbook_has_a_methodology_sheet(db_session):
    _seed(db_session)
    assert "Methodology" in _workbook(db_session).sheetnames


def test_methodology_lists_every_rendered_section_with_finding_and_disclosure(db_session):
    _seed(db_session)
    wb = _workbook(db_session)
    rows = _rows(wb["Methodology"])

    # Header plus one row per section sheet the workbook actually wrote.
    listed = {row[0] for row in rows[1:]}
    assert "Entity Statistics" in listed
    assert "Journal Portfolio" in listed

    by_sheet = {row[0]: row for row in rows[1:]}
    journal = by_sheet["Journal Portfolio"]
    assert "NIF" in journal[1] or "journal" in journal[1].lower(), journal[1]
    # The disclosure names what the figure is not — the reason this sheet exists.
    assert "not the journal impact factor" in journal[2].lower(), journal[2]


def test_every_sheet_methodology_names_exists_in_the_workbook(db_session):
    """A disclosure pointing at a sheet that is not there is a dead reference."""
    _seed(db_session)
    wb = _workbook(db_session)
    for row in _rows(wb["Methodology"])[1:]:
        assert row[0] in wb.sheetnames, f"Methodology cites missing sheet {row[0]!r}"


def test_methodology_does_not_invent_exhibit_ordinals(db_session):
    """Design decision 7. Excel renders a different set than the document, so a
    workbook numbering its own exhibits would disagree with the PDF of the same
    generation and say nothing about it. The sheet name is the reference."""
    _seed(db_session)
    blob = "\n".join(
        str(cell.value)
        for row in _workbook(db_session)["Methodology"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Exhibit" not in blob


def test_methodology_omits_sections_that_were_not_requested(db_session):
    _seed(db_session)
    wb = _workbook(db_session, sections=["entity_stats"])
    listed = {row[0] for row in _rows(wb["Methodology"])[1:]}
    assert "Entity Statistics" in listed
    assert "Journal Portfolio" not in listed


# ── 5.3 Caveat travels with the range ─────────────────────────────────────────


def test_the_bespoke_harmonization_sheet_is_covered_too(db_session):
    """The disclosure follows what a format *renders*, not how it renders it.

    Harmonization still comes from a bespoke writer rather than the shared
    payload — its sheet carries row ids, executed-at and reverted over up to 200
    rows, detail the collector's table does not have. That makes it the one
    section whose sheet could quietly sit outside the contract while the parity
    map claims Excel renders it.
    """
    _seed(db_session)
    wb = _workbook(db_session)

    listed = {row[0] for row in _rows(wb["Methodology"])[1:]}
    assert "Harmonization" in listed

    ws = wb["Harmonization"]
    # Same shape every migrated sheet has: finding, caveat, then the header.
    assert "harmonization operation" in str(ws["A1"].value), ws["A1"].value
    assert "were applied, not proposed" in str(ws["A2"].value), ws["A2"].value
    assert ws["A3"].value == "ID", "the header should sit below both"
    # Freezing the caveat instead of the header would scroll the columns away.
    assert ws.freeze_panes == "A4"


#: The plural nouns these sections count. Spelled out rather than matched by
#: pattern because English will not cooperate: `1 patterns` is a defect and
#: `1 remains unresolved` is correct, and a plural noun is spelled exactly like a
#: third-person-singular verb. The first version of this test read "remains" as a
#: noun and failed on a correct sentence.
_COUNTED_PLURALS = (
    "entities", "authority records", "authors", "communities", "collaborations",
    "journals", "patterns", "recommended actions", "harmonization operations",
    "rules", "concepts",
)


def test_no_takeaway_counts_one_of_something_in_the_plural(db_session):
    """A takeaway is a heading in HTML and the first row of a sheet in Excel, so
    `1 patterns detected` is the section's most prominent line rather than one
    entry in a summary list.

    Seeded deliberately singular — one entity, one authority record, one
    operation — because that is the only dataset where this can go wrong.
    """
    import re

    from backend import report_builder

    _seed_singular(db_session)
    every_section = [s for s in report_builder.SECTION_COLLECTORS if s != "top_brands"]
    wb = _workbook(db_session, sections=every_section)
    findings = [row[1] for row in _rows(wb["Methodology"])[1:]]
    assert len(findings) >= 10, "the sweep has to cover every counted noun"

    for finding in findings:
        for noun in _COUNTED_PLURALS:
            # `(?<![\d,])` so "11 entities" and "21 entities" are not read as
            # a count of one.
            assert not re.search(rf"(?<![\d,])1 {noun}\b", finding), (
                f"{noun!r} is plural after a count of one: {finding!r}"
            )


def test_the_journal_table_carries_its_caveat_on_the_row_above(db_session):
    """The case the requirement was written for: a NIF column pasted into a deck
    without the sentence saying it is not the Journal Impact Factor."""
    _seed(db_session)
    ws = _workbook(db_session)["Journal Portfolio"]

    # Match the column heading itself, not any cell mentioning NIF — the caveat
    # names the metric too, which is the whole point of it.
    header_row = next(
        cell.row
        for row in ws.iter_rows()
        for cell in row
        if cell.value == "NIF (field-normalized)"
    )
    above = ws.cell(row=header_row - 1, column=1).value
    assert above, f"no caveat row above the table header at row {header_row}"
    assert "not the journal impact factor" in above.lower(), above
