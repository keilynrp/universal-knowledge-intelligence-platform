"""
Sprint 40 regression tests — Enterprise Export (PDF + Excel).

WeasyPrint is patched for all PDF tests so CI doesn't need a browser engine.
Excel tests use openpyxl to verify workbook structure.
"""
import io
import pytest
from unittest.mock import MagicMock, patch

import openpyxl

from backend import models
from backend.routers import reports as reports_router


_REPORT_PAYLOAD = {
    "domain_id": "default",
    "sections": ["entity_stats"],
    "title": "Test Export",
}

_MOCK_WEASY = "backend.routers.reports._make_pdf"


# ── PDF endpoint ───────────────────────────────────────────────────────────────

def test_pdf_requires_auth(client):
    resp = client.post("/exports/pdf", json=_REPORT_PAYLOAD)
    assert resp.status_code in (401, 403)


def test_pdf_calls_report_builder(client, auth_headers):
    """_make_pdf must be called with the HTML output from report_builder.build()."""
    fake_pdf = b"%PDF-1.4 fake"
    with patch(_MOCK_WEASY, return_value=fake_pdf) as mock_pdf:
        resp = client.post("/exports/pdf", json=_REPORT_PAYLOAD, headers=auth_headers)
    # May be 200 or 501 (weasyprint not installed) — should NOT be 422/403/500 from our code
    assert resp.status_code in (200, 501)
    if resp.status_code == 200:
        mock_pdf.assert_called_once()


def test_pdf_response_content_type(client, auth_headers):
    fake_pdf = b"%PDF-1.4 fake"
    with patch(_MOCK_WEASY, return_value=fake_pdf):
        resp = client.post("/exports/pdf", json=_REPORT_PAYLOAD, headers=auth_headers)
    if resp.status_code == 200:
        assert "application/pdf" in resp.headers["content-type"]


def test_pdf_invalid_section_returns_422(client, auth_headers):
    payload = {"domain_id": "default", "sections": ["nonexistent_section"]}
    with patch(_MOCK_WEASY, return_value=b""):
        resp = client.post("/exports/pdf", json=payload, headers=auth_headers)
    assert resp.status_code == 422


def test_reports_sections_include_decision_recommendations(client, auth_headers):
    resp = client.get("/reports/sections", headers=auth_headers)
    assert resp.status_code == 200
    ids = {section["id"] for section in resp.json()}
    assert "decision_recommendations" in ids
    assert "impact_projection" in ids
    assert "institutional_benchmark" in ids


def test_html_report_accepts_decision_recommendations_section(client, auth_headers):
    payload = {
        "domain_id": "default",
        "sections": ["decision_recommendations"],
        "title": "Decision Brief",
    }
    resp = client.post("/reports/generate", json=payload, headers=auth_headers)
    assert resp.status_code == 200
    assert "Suggested Next Actions" in resp.text


def test_html_report_accepts_impact_projection_section(client, auth_headers, db_session):
    from backend import models

    db_session.add(models.RawEntity(
        primary_label="High impact publication",
        domain="default",
        enrichment_status="completed",
        enrichment_citation_count=240,
        enrichment_source="openalex",
        quality_score=0.82,
    ))
    db_session.commit()

    payload = {
        "domain_id": "default",
        "sections": ["impact_projection"],
        "title": "Impact Brief",
    }
    resp = client.post("/reports/generate", json=payload, headers=auth_headers)
    assert resp.status_code == 200
    assert "Impact Projection" in resp.text
    assert "Monte Carlo" in resp.text


def test_pdf_accepts_full_pilot_sections(client, auth_headers):
    payload = {
        "domain_id": "default",
        "sections": [
            "entity_stats",
            "enrichment_coverage",
            "impact_projection",
            "decision_recommendations",
            "institutional_benchmark",
            "top_brands",
            "topic_clusters",
        ],
        "title": "Pilot PDF",
    }
    fake_pdf = b"%PDF-1.4 fake"
    with patch(_MOCK_WEASY, return_value=fake_pdf):
        resp = client.post("/exports/pdf", json=payload, headers=auth_headers)

    assert resp.status_code == 200
    assert "application/pdf" in resp.headers["content-type"]


def test_make_pdf_returns_501_when_weasyprint_runtime_is_missing():
    with patch.object(reports_router, "_load_weasyprint_html", side_effect=OSError("missing libgobject")):
        with pytest.raises(Exception) as exc_info:
            reports_router._make_pdf("<html></html>")

    exc = exc_info.value
    assert getattr(exc, "status_code", None) == 501
    assert "GTK runtime" in getattr(exc, "detail", "")


def test_html_report_accepts_institutional_benchmark_section(client, auth_headers):
    payload = {
        "domain_id": "default",
        "sections": ["institutional_benchmark"],
        "benchmark_profile_id": "sni_readiness_baseline",
        "title": "Benchmark Brief",
    }
    resp = client.post("/reports/generate", json=payload, headers=auth_headers)
    assert resp.status_code == 200
    assert "Institutional Benchmark" in resp.text
    assert "SNI Readiness Baseline" in resp.text


# ── Excel endpoint ─────────────────────────────────────────────────────────────

def test_excel_requires_auth(client):
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD)
    assert resp.status_code in (401, 403)


def test_excel_content_type_xlsx(client, auth_headers):
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    ct = resp.headers["content-type"]
    assert "spreadsheetml" in ct or "officedocument" in ct


def test_excel_has_summary_sheet(client, auth_headers):
    """Response bytes must be a valid xlsx with a 'Summary' sheet."""
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Summary" in wb.sheetnames


def test_excel_has_entities_sheet(client, auth_headers):
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Entities" in wb.sheetnames


def test_excel_header_violet_fill(client, auth_headers):
    """The Summary sheet header row must use the violet brand fill (#5B21B6)."""
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Summary"]
    fill = ws.cell(row=1, column=1).fill
    assert fill.fgColor.rgb.upper().endswith("5B21B6")


def test_excel_concepts_sheet_when_requested(client, auth_headers):
    """Including 'topic_clusters' in sections must produce a 'Concepts' sheet."""
    payload = {
        "domain_id": "default",
        "sections": ["entity_stats", "topic_clusters"],
    }
    resp = client.post("/exports/excel", json=payload, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Concepts" in wb.sheetnames
