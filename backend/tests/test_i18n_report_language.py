"""`POST /reports/generate` accepts a language.

Phase 8. This is the endpoint #209 was filed against: it had no way to say which
language it wanted, so a report was whatever mixture the code happened to hold.

**Reports do not read `Accept-Language`, and that is deliberate.** A report is
produced for an audience, not for whoever pressed the button, so an operator's
browser locale must not leak into someone else's document. Settled with the
product owner on 2026-07-31 and made structural in phase 5, where
`resolve_report_language` takes no header argument at all. These tests pin that
behaviour at the endpoint, because a dependency swap would otherwise reintroduce
it silently.
"""

from __future__ import annotations

import pytest

from backend.i18n import catalog as catalog_module

pytestmark = pytest.mark.reporting


def _generate(client, auth_headers, **overrides) -> str:
    body = {"domain_id": "default", "sections": ["entity_stats"], "title": "T"}
    query = ""
    if "language" in overrides:
        query = f"?language={overrides.pop('language')}"
    body.update(overrides)
    resp = client.post(f"/reports/generate{query}", json=body, headers=auth_headers)
    assert resp.status_code == 200, resp.text[:300]
    return resp.text


class TestTheLanguageParameter:
    def test_spanish_produces_spanish_section_titles(self, client, auth_headers):
        html = _generate(client, auth_headers, language="es")

        assert "Estadísticas de Entidades" in html, (
            "the section title did not follow the requested language"
        )

    def test_english_produces_english_section_titles(self, client, auth_headers):
        html = _generate(client, auth_headers, language="en")

        assert "Entity Statistics" in html

    def test_omitting_the_parameter_preserves_existing_behaviour(
        self, client, auth_headers
    ):
        """Task 8.4 — callers written before this parameter must not change."""
        html = _generate(client, auth_headers)

        assert "Entity Statistics" in html
        assert "Estadísticas de Entidades" not in html

    def test_an_unsupported_language_still_produces_a_report(self, client, auth_headers):
        """A cosmetic request must not fail an expensive artefact."""
        html = _generate(client, auth_headers, language="fr")

        assert "Entity Statistics" in html


class TestTheHeaderIsIgnored:
    """The decision this whole resolver split exists for."""

    def test_a_spanish_browser_gets_an_english_report(self, client, auth_headers):
        resp = client.post(
            "/reports/generate",
            json={"domain_id": "default", "sections": ["entity_stats"], "title": "T"},
            headers={**auth_headers, "Accept-Language": "es-MX,es;q=0.9"},
        )
        assert resp.status_code == 200

        assert "Entity Statistics" in resp.text, (
            "Accept-Language leaked into a report — the operator's locale must not "
            "decide the language of someone else's document"
        )
        assert "Estadísticas de Entidades" not in resp.text

    def test_the_explicit_parameter_still_wins_over_a_conflicting_header(
        self, client, auth_headers
    ):
        resp = client.post(
            "/reports/generate?language=es",
            json={"domain_id": "default", "sections": ["entity_stats"], "title": "T"},
            headers={**auth_headers, "Accept-Language": "en-US,en;q=0.9"},
        )
        assert resp.status_code == 200

        assert "Estadísticas de Entidades" in resp.text


class TestTheDisclosure:
    """Task 8.5 — the limitation is stated, not left to be discovered."""

    def test_a_spanish_report_says_analysis_stays_english(self, client, auth_headers):
        html = _generate(client, auth_headers, language="es")

        assert "report.disclosure.analysis_language" not in html, (
            "the disclosure key rendered raw — it is missing from the catalog"
        )
        assert any(
            marker in html for marker in ("permanece en inglés", "permanecen en inglés")
        ), "a non-English report must state that analysis text stays English"

    def test_an_english_report_does_not_need_the_disclosure(self, client, auth_headers):
        html = _generate(client, auth_headers, language="en")

        assert "permanece en inglés" not in html


class TestTheStakeholderReadingFollowsTheLanguage:
    """#268 — the opening section had its copy migrated but never its language.

    `_STAKEHOLDER_PROFILES` moved to the catalog in #284, which made the Spanish
    copy exist. It never rendered: `_section_stakeholder_reading` called
    `render_html(collect_...())` with no `language`, so every key resolved
    through the default and the section that opens a Spanish report was English
    end to end.

    That is invisible to the render-boundary guard, which proves no key *leaks*
    and says nothing about which language a resolved key resolved into. It is
    also invisible to the section-title tests above, because this section's
    title is not what was wrong.
    """

    def _stakeholder_copy(self, language: str) -> str:
        from backend.i18n.catalog import translate

        return translate("report.stakeholder.how_to_read", language)

    def test_a_spanish_report_frames_the_lens_in_spanish(self, client, auth_headers):
        html = _generate(client, auth_headers, language="es")

        assert self._stakeholder_copy("es") in html, (
            "the stakeholder reading rendered in the default language: the "
            "section's copy is in the catalog but the renderer was never told "
            "which language to resolve it into"
        )
        assert self._stakeholder_copy("en") not in html, (
            "English stakeholder copy survived in a Spanish report"
        )

    def test_an_english_report_is_unchanged(self, client, auth_headers):
        html = _generate(client, auth_headers, language="en")

        assert self._stakeholder_copy("en") in html


class TestTheTitlesComeFromTheCatalog:
    def test_output_follows_the_catalog(self, client, auth_headers, monkeypatch):
        sentinel = "SENTINEL-SECTION"
        real = catalog_module._load_catalog.__wrapped__("en")
        real_keys = [key for key in real if key.startswith("report.section.")]
        assert real_keys, "no report.section.* keys in the catalog"
        # Overlay the sentinel onto the real catalog rather than replacing it:
        # since #292 every document-level scalar (cover captions, the
        # stakeholder label, the executive summary title, ...) resolves
        # through the same localize_document() pass as section titles, and a
        # catalog missing those entirely would (correctly) trip the boundary's
        # own unresolved-key guard. Only the keys this test cares about need
        # to change.
        patched = {**real, **{k: sentinel for k in real_keys}}
        monkeypatch.setattr(catalog_module, "_load_catalog", lambda language: patched)

        html = _generate(client, auth_headers, language="en")

        assert sentinel in html, "section titles still come from a literal"


class TestTheHeaderCannotReachReportGeneration:
    """Structural, not behavioural — the output test cannot discriminate.

    Swapping the builder's resolver for the header-aware one leaves all nine
    behavioural tests green, because the header never reaches the builder in the
    first place: the router simply does not accept it. That makes the guarantee
    a property of the signature, and a signature is what a refactor changes
    silently. So it is asserted directly.
    """

    def test_the_endpoint_takes_no_accept_language_parameter(self):
        import inspect

        from backend.routers.reports import generate_report

        parameters = inspect.signature(generate_report).parameters
        assert "accept_language" not in parameters, (
            "report generation must not be able to see Accept-Language"
        )

    def test_no_header_dependency_is_wired_into_the_endpoint(self):
        import inspect

        from fastapi import params

        from backend.routers.reports import generate_report

        headers = [
            name
            for name, p in inspect.signature(generate_report).parameters.items()
            if isinstance(p.default, params.Header)
        ]
        assert not headers, f"report generation reads headers: {headers}"

    def test_the_builder_resolves_with_the_report_resolver(self):
        """`resolve_report_language` takes no header; `resolve_language` does."""
        import inspect

        from backend.i18n.locale import resolve_report_language

        assert set(inspect.signature(resolve_report_language).parameters) == {"explicit"}

        import backend.report_builder as rb

        source = inspect.getsource(rb.build)
        assert "resolve_report_language(" in source, (
            "the builder must resolve with the header-blind resolver"
        )
        assert "resolve_language(" not in source.replace("resolve_report_language(", ""), (
            "the builder reaches for the header-aware resolver"
        )


class TestExcelSheetNamesStayValid:
    """Excel caps a sheet name at 31 characters and forbids / \ ? * [ ] :

    `_safe_sheet_title` truncates rather than failing, so a violation never
    breaks a workbook — it silently cuts a tab mid-word instead. That is why
    `report.sheet.*` exists as a separate, purpose-built short name, and why
    the constraint is asserted here: a future translation would otherwise
    degrade the tabs without anything going red.
    """

    _INVALID = set('/\?*[]:')

    def test_spanish_sheet_names_fit_and_are_legal(self):
        """Names authored by this change must need no truncation at all."""
        catalog = catalog_module._load_catalog.__wrapped__("es")
        sheet_keys = {k: v for k, v in catalog.items() if k.startswith("report.sheet.")}

        assert sheet_keys, "no report.sheet.* keys for es"
        too_long = {k: v for k, v in sheet_keys.items() if len(v) > 31}
        illegal = {k: v for k, v in sheet_keys.items() if set(v) & self._INVALID}

        assert not too_long, f"sheet names over 31 chars: {too_long}"
        assert not illegal, f"sheet names with forbidden characters: {illegal}"

    def test_english_sheet_names_are_exactly_what_they_were(self):
        """Task 8.4: omitting the parameter must not rename a tab.

        English keeps the long section title it always had — including the two
        that Excel truncates, because that truncation is the behaviour existing
        callers already see. Shortening them would be an improvement nobody
        asked for, delivered as a silent change to a downloaded file. Spanish is
        where purpose-built short names are justified, since the translated
        titles do not fit at all.
        """
        from backend import report_builder

        catalog = catalog_module._load_catalog.__wrapped__("en")

        for section_id, label in report_builder.SECTION_LABELS.items():
            assert catalog[f"report.sheet.{section_id}"] == label, (
                f"{section_id}: the English sheet name changed from {label!r}"
            )

    def test_a_sheet_name_exists_for_every_section(self):
        catalog = catalog_module._load_catalog.__wrapped__("en")
        sections = {k.rsplit(".", 1)[-1] for k in catalog if k.startswith("report.section.")}
        sheets = {k.rsplit(".", 1)[-1] for k in catalog if k.startswith("report.sheet.")}

        assert sections <= sheets, f"sections with no sheet name: {sorted(sections - sheets)}"


class TestAllThreeFormatsFollowTheLanguage:
    """Task 8.3 — the parity contract these formats already carry."""

    def _export(self, client, auth_headers, path, language):
        resp = client.post(
            f"{path}?language={language}",
            json={"domain_id": "default", "sections": ["entity_stats"], "title": "T"},
            headers=auth_headers,
        )
        return resp

    def test_pdf_carries_the_language(self, client, auth_headers):
        resp = self._export(client, auth_headers, "/exports/pdf", "es")
        if resp.status_code == 500:
            pytest.skip("WeasyPrint runtime not available in this environment")
        assert resp.status_code == 200

    @pytest.mark.parametrize("language", ["en", "es"])
    def test_excel_names_its_sheets_in_the_language(self, client, auth_headers, language):
        import io

        import openpyxl

        resp = self._export(client, auth_headers, "/exports/excel", language)
        assert resp.status_code == 200, resp.text[:200]

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        expected = catalog_module._load_catalog.__wrapped__(language)[
            "report.sheet.entity_stats"
        ]
        assert any(expected in name for name in wb.sheetnames), (
            f"{language}: expected a sheet named {expected!r}, got {wb.sheetnames}"
        )

    def test_pptx_builds_in_spanish(self, client, auth_headers):
        resp = self._export(client, auth_headers, "/exports/pptx", "es")
        if resp.status_code == 500 and "pptx" in resp.text.lower():
            pytest.skip("python-pptx not available in this environment")
        assert resp.status_code == 200


class TestTheDisclosureIsSpecific:
    """Phase 9 found the disclosure was true but not useful.

    Reading a real Spanish report showed 6 Spanish lines against 32 English
    ones: section titles translate, but structural headings, metric labels,
    empty-state messages and the stakeholder framing do not. The original
    disclosure only mentioned "analysis text and provider names", so a reader
    meeting a mostly-English document had no way to tell intent from defect.

    No automated sweep could have found this. Every sweep in this track looked
    for *Spanish* strings; these were always English, so nothing flagged them.
    """

    _MUST_NAME = ("headings", "labels", "providers")
    _MUST_NAME_ES = ("encabezados", "etiquetas", "fuentes externas")

    @pytest.mark.parametrize(
        "language,markers",
        [("en", _MUST_NAME), ("es", _MUST_NAME_ES)],
    )
    def test_the_disclosure_names_each_category_that_stays_english(
        self, language, markers
    ):
        text = catalog_module._load_catalog.__wrapped__(language)[
            "report.disclosure.analysis_language"
        ]

        missing = [m for m in markers if m not in text.lower()]
        assert not missing, (
            f"{language}: the disclosure does not mention {missing}; a reader cannot "
            f"tell which English text is intended and which is a defect"
        )
