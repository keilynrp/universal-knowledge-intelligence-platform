"""
Enterprise Excel exporter — multi-sheet branded workbook.
Used by POST /exports/excel in backend/routers/reports.py.
"""
from __future__ import annotations

import logging
from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend import models
from backend.analyzers.topic_modeling import TopicAnalyzer
from backend.tenant_access import scope_query_to_org

logger = logging.getLogger(__name__)

# Violet brand palette (hex, no #)
_HEADER_FG   = "5B21B6"   # violet-800
_HEADER_FONT = "FFFFFF"   # white

_HEADER_FILL = PatternFill("solid", fgColor=_HEADER_FG)
_HEADER_FONT_STYLE = Font(color=_HEADER_FONT, bold=True, size=11)
_SUBHEADER_FONT = Font(bold=True, size=10)


def _style_header_row(ws, cols: list[str]) -> None:
    """Write and style a header row (row 1) with violet fill + white bold text."""
    for col_idx, header in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT_STYLE
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Approximate column width based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


class EnterpriseExcelExporter:
    """Build a branded multi-sheet Excel workbook and return raw bytes."""

    _ENTITY_CAP = 5_000
    _CONCEPT_CAP = 50

    def _entities_query(self, db: Session, domain_id: str, org_id: int | None):
        query = scope_query_to_org(db.query(models.RawEntity), models.RawEntity, org_id)
        if domain_id:
            query = query.filter(models.RawEntity.domain == domain_id)
        return query

    def _harmonization_query(self, db: Session, org_id: int | None):
        return scope_query_to_org(db.query(models.HarmonizationLog), models.HarmonizationLog, org_id)

    def build(
        self,
        db: Session,
        domain_id: str,
        sections: List[str],
        org_id: int | None = None,
        manual_sections: list[dict[str, str]] | None = None,
    ) -> bytes:
        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary KPIs ──────────────────────────────────────────────
        ws_summary = wb.active
        assert ws_summary is not None
        ws_summary.title = "Summary"
        self._write_summary(ws_summary, db, domain_id, org_id)

        # ── Sheet 2: Entities ──────────────────────────────────────────────────
        ws_entities = wb.create_sheet("Entities")
        self._write_entities(ws_entities, db, domain_id, org_id)

        # ── Sheet 3: Concepts ─────────────────────────────────────────────────
        if "topic_clusters" in sections:
            ws_concepts = wb.create_sheet("Concepts")
            self._write_concepts(ws_concepts, db, domain_id, org_id)

        # ── Migrated sections: rendered from the shared section payload ───────
        # These render via the format-neutral collector + Excel renderer, so the
        # section is authored once and appears here without a bespoke writer.
        # (unify-report-format-coverage phase 3; entity_stats is the pilot.)
        if "entity_stats" in sections:
            from backend import report_builder
            from backend.reporting.excel_renderer import render_excel
            render_excel(report_builder.collect_entity_stats(db, domain_id, org_id), wb)

        # ── Sheet 4: Harmonization Log ────────────────────────────────────────
        if "harmonization_log" in sections:
            ws_harm = wb.create_sheet("Harmonization")
            self._write_harmonization(ws_harm, db, org_id)

        if manual_sections:
            ws_notes = wb.create_sheet("Analyst Notes")
            self._write_manual_sections(ws_notes, manual_sections)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Private sheet writers ──────────────────────────────────────────────────

    def _write_summary(self, ws, db: Session, domain_id: str, org_id: int | None) -> None:
        query = self._entities_query(db, domain_id, org_id)
        total = query.count()
        enriched = (
            query
            .filter(models.RawEntity.enrichment_status == "completed")
            .count()
        )
        pct = round(enriched / total * 100, 1) if total > 0 else 0.0

        from sqlalchemy import func
        avg_row = (
            query.with_entities(func.avg(models.RawEntity.enrichment_citation_count))
            .filter(models.RawEntity.enrichment_status == "completed")
            .scalar()
        )
        avg_cit = round(float(avg_row), 1) if avg_row is not None else 0.0

        headers = ["Metric", "Value"]
        _style_header_row(ws, headers)

        rows = [
            ("Active Domain",       domain_id),
            ("Total Entities",      total),
            ("Enriched Entities",   enriched),
            ("Enrichment %",        f"{pct}%"),
            ("Avg Citations",       avg_cit),
            ("Platform",            "UKIP — Universal Knowledge Intelligence Platform"),
        ]
        for row_idx, (metric, value) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=metric).font = _SUBHEADER_FONT
            ws.cell(row=row_idx, column=2, value=value)

        _autofit(ws)

    def _write_entities(self, ws, db: Session, domain_id: str, org_id: int | None) -> None:
        headers = [
            "ID", "Primary Label", "Secondary Label", "Canonical ID", "Entity Type",
            "Enrichment Status", "Citation Count", "Source",
        ]
        _style_header_row(ws, headers)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        rows = (
            self._entities_query(db, domain_id, org_id)
            .order_by(models.RawEntity.id)
            .limit(self._ENTITY_CAP)
            .all()
        )
        for row_idx, e in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=e.id)
            ws.cell(row=row_idx, column=2, value=e.primary_label)
            ws.cell(row=row_idx, column=3, value=e.secondary_label)
            ws.cell(row=row_idx, column=4, value=e.canonical_id)
            ws.cell(row=row_idx, column=5, value=e.entity_type)
            ws.cell(row=row_idx, column=6, value=e.enrichment_status)
            ws.cell(row=row_idx, column=7, value=e.enrichment_citation_count)
            ws.cell(row=row_idx, column=8, value=e.enrichment_source)

        _autofit(ws)

    def _write_concepts(
        self,
        ws,
        db: Session,
        domain_id: str,
        org_id: int | None,
    ) -> None:  # db kept for consistency
        headers = ["Rank", "Concept", "Count", "Percentage (%)"]
        _style_header_row(ws, headers)

        try:
            result = TopicAnalyzer().top_topics(
                domain_id=domain_id,
                top_n=self._CONCEPT_CAP,
                org_id=org_id,
            )
            topics = result.get("topics", [])
        except Exception:
            logger.warning("TopicAnalyzer failed in excel_exporter", exc_info=True)
            topics = []

        for row_idx, t in enumerate(topics, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=t.get("concept", ""))
            ws.cell(row=row_idx, column=3, value=t.get("count", 0))
            ws.cell(row=row_idx, column=4, value=round(t.get("pct", 0.0), 2))

        _autofit(ws)

    def _write_harmonization(self, ws, db: Session, org_id: int | None) -> None:
        headers = ["ID", "Step ID", "Step Name", "Records Updated", "Fields Modified", "Executed At", "Reverted"]
        _style_header_row(ws, headers)

        rows = (
            self._harmonization_query(db, org_id)
            .order_by(models.HarmonizationLog.id.desc())
            .limit(200)
            .all()
        )
        for row_idx, h in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=h.id)
            ws.cell(row=row_idx, column=2, value=h.step_id)
            ws.cell(row=row_idx, column=3, value=h.step_name)
            ws.cell(row=row_idx, column=4, value=h.records_updated)
            ws.cell(row=row_idx, column=5, value=h.fields_modified)
            ws.cell(row=row_idx, column=6, value=str(h.executed_at) if h.executed_at else "")
            ws.cell(row=row_idx, column=7, value="Yes" if h.reverted else "No")

        _autofit(ws)

    def _write_manual_sections(self, ws, manual_sections: list[dict[str, str]]) -> None:
        headers = ["Section", "Analyst Text"]
        _style_header_row(ws, headers)
        for row_idx, section in enumerate(manual_sections, start=2):
            ws.cell(row=row_idx, column=1, value=(section.get("title") or "Analyst Note")[:120])
            cell = ws.cell(row=row_idx, column=2, value=section.get("content") or "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 90
